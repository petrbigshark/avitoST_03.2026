import argparse
import os
import re
import subprocess
import sys
import tempfile
import time
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple
from urllib.parse import quote_plus

import pandas as pd
import requests
from openpyxl import load_workbook

sys.dont_write_bytecode = True


IMGBB_API_KEY = os.getenv("IMGBB_API_KEY", "").strip()

PROJECT_DIR = Path(__file__).resolve().parent
PHOTOS_ROOT = PROJECT_DIR / "Avito_ST_2"
OUTPUT_ROOT_DIR = PROJECT_DIR / "output Готовые файлы"
OUTPUT_ROOT_DIR.mkdir(parents=True, exist_ok=True)

TARGET_SHEET = "Женская обувь-Кроссовки и кеды"
ID_HEADER = "Уникальный идентификатор объявления"
NAME_HEADER = "Название объявления"
DESCRIPTION_HEADER = "Описание объявления"
SIZE_HEADER = "Размер"
PHOTOS_HEADER = "Ссылки на фото"
PRICE_HEADER = "Цена"
CATEGORY_HEADER = "Категория"
GOODS_TYPE_HEADER = "Вид одежды"
CONDITION_HEADER = "Состояние"
BRAND_HEADER = "Бренд одежды"
COLOR_HEADER = "Цвет"
COLOR_MANUFACTURER_HEADER = "Цвет от производителя"
APPAREL_TYPE_HEADER = "Вид одежды, обуви, аксессуаров"
MULTI_TITLE_HEADER = "Название мультиобъявления"
TYPE_GOODS_HEADER = "Тип товара"
TYPE_SHOES_HEADER = "Тип одежды или обуви"
DEFAULT_CATEGORY_VALUE = "Одежда, обувь, аксессуары"
KIDS_TYPE_GOODS_VALUE = "Обувь"
KIDS_TYPE_SHOES_VALUE = "Кроссовки"

BASE_FIELDS = ("GoodsType", "Color", "ApparelType")
PHOTO_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".bmp", ".gif", ".heic", ".heif"}
HEIC_EXTENSIONS = {".heic", ".heif"}
MULTI_TITLE_ALLOWED_CHARS_RE = re.compile(r"[^A-Za-zА-Яа-яЁё0-9\-/ ]+")
IMGBB_KEY_RE = re.compile(r"(key=)[^&\s]+", re.IGNORECASE)


def normalize_id(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if re.fullmatch(r"\d+\.0+", text):
        return text.split(".", 1)[0]
    return text


def normalize_article(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = text.replace("—", "-").replace("–", "-")
    text = re.sub(r"\s+", " ", text)
    return text


def normalize_human_case(value: object) -> str:
    """
    Normalizes free text into consistent title-like casing:
    adidas Y-3 GAZELLE white -> Adidas Y-3 Gazelle White
    """
    if value is None:
        return ""

    text = str(value).strip()
    if not text:
        return ""
    text = re.sub(r"\s+", " ", text)

    def normalize_chunk(chunk: str) -> str:
        if not chunk:
            return chunk
        lower = chunk.lower()
        match = re.search(r"[A-Za-zА-Яа-яЁё]", lower)
        if not match:
            return lower
        pos = match.start()
        return lower[:pos] + lower[pos].upper() + lower[pos + 1 :]

    def normalize_token(token: str) -> str:
        slash_parts = token.split("/")
        norm_slash = []
        for slash_part in slash_parts:
            hyphen_parts = slash_part.split("-")
            norm_hyphen = [normalize_chunk(part) for part in hyphen_parts]
            norm_slash.append("-".join(norm_hyphen))
        return "/".join(norm_slash)

    return " ".join(normalize_token(token) for token in text.split(" "))


def normalize_brand(value: object) -> str:
    normalized = normalize_human_case(value)
    if not normalized:
        return ""
    key = article_key(normalized)
    if "NIKE" in key and "OFFWHITE" in key:
        return "Nike"
    return normalized


def add_original_suffix(title: object) -> str:
    normalized = clean_text(title)
    if not normalized:
        return ""
    if re.search(r"оригинал\s*$", normalized, flags=re.IGNORECASE):
        return normalized
    return f"{normalized} Оригинал"


def sanitize_multi_title(value: object) -> str:
    text = clean_text(value)
    if not text:
        return ""
    text = text.replace("—", "-").replace("–", "-")
    text = MULTI_TITLE_ALLOWED_CHARS_RE.sub(" ", text)
    return re.sub(r"\s+", " ", text).strip()


def mask_sensitive_text(value: object) -> str:
    text = clean_text(value)
    if not text:
        return ""
    return IMGBB_KEY_RE.sub(r"\1***", text)


def article_key(value: object) -> str:
    text = normalize_article(value).upper()
    text = text.replace("Ё", "Е")
    return re.sub(r"[^A-Z0-9А-Я]+", "", text)


def format_size_value(value: object) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return f"{value:.1f}".rstrip("0").rstrip(".").replace(".", ",")
    text = str(value).strip()
    if re.fullmatch(r"\d+\.0+", text):
        text = text.split(".", 1)[0]
    else:
        text = text.replace(".", ",")
    return text


def parse_price(value: object) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    text = str(value)
    digits = re.sub(r"[^\d]", "", text)
    return digits


def clean_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    text = str(value).strip()
    if text.lower() == "nan":
        return ""
    return text


def get_discount_price_from_avitonow(row: pd.Series) -> str:
    # Supports both historical and new column names in AvitoNow.
    return parse_price(row.get("Цена после скидки") or row.get("Цена со скидкой"))


def fold_text(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value).lower()
    return "".join(char for char in normalized if not unicodedata.combining(char))


def find_input_file(patterns: Iterable[str]) -> Path:
    files = [
        p for p in PROJECT_DIR.iterdir()
        if p.is_file() and not p.name.startswith("~$") and p.suffix.lower() == ".xlsx"
    ]
    for pattern in patterns:
        regex = re.compile(pattern, re.IGNORECASE)
        folded_regex = re.compile(fold_text(pattern), re.IGNORECASE)
        for file_path in files:
            file_name = file_path.name
            if regex.search(file_name):
                return file_path
            if folded_regex.search(fold_text(file_name)):
                return file_path
    raise FileNotFoundError(f"Input file not found by patterns: {patterns}")


def choose_next_version(date_str: str) -> int:
    folder_pattern = re.compile(rf"^Готовые файлы {re.escape(date_str)} v(\d+)$")
    file_pattern = re.compile(rf"^Готовый файл {re.escape(date_str)} v(\d+)\.xlsx$")
    versions: List[int] = []

    for path in OUTPUT_ROOT_DIR.iterdir():
        folder_match = folder_pattern.match(path.name) if path.is_dir() else None
        file_match = file_pattern.match(path.name) if path.is_file() else None
        match = folder_match or file_match
        if match:
            versions.append(int(match.group(1)))

    return (max(versions) + 1) if versions else 1


def parse_version_arg(raw_value: object) -> int:
    text = clean_text(raw_value).lower()
    if not text:
        raise ValueError("Version is empty")
    if text.startswith("v"):
        text = text[1:]
    if not text.isdigit():
        raise ValueError(f"Invalid version value '{raw_value}'. Expected vN or N.")
    parsed = int(text)
    if parsed <= 0:
        raise ValueError(f"Version must be > 0, got '{raw_value}'.")
    return parsed


def migrate_legacy_output_files() -> List[str]:
    """
    Moves legacy flat output files into versioned run folders.
    Example:
    - output Готовые файлы/Готовый файл 01.03.2026 v5.xlsx
    ->
    - output Готовые файлы/Готовые файлы 01.03.2026 v5/Готовый файл 01.03.2026 v5.xlsx
    """
    patterns = [
        re.compile(r"^(Готовый файл) (\d{2}\.\d{2}\.\d{4}) v(\d+)\.xlsx$"),
        re.compile(r"^(IMGBB ссылки) (\d{2}\.\d{2}\.\d{4}) v(\d+)\.xlsx$"),
        re.compile(r"^(Лог) (\d{2}\.\d{2}\.\d{4}) v(\d+)\.txt$"),
    ]
    moved_messages: List[str] = []

    for path in OUTPUT_ROOT_DIR.iterdir():
        if not path.is_file() or path.name.startswith("~$"):
            continue

        matched = None
        for pattern in patterns:
            match = pattern.match(path.name)
            if match:
                matched = match
                break
        if matched is None:
            continue

        _, date_str, version = matched.groups()
        run_dir = OUTPUT_ROOT_DIR / f"Готовые файлы {date_str} v{version}"
        run_dir.mkdir(parents=True, exist_ok=True)
        target_path = run_dir / path.name

        if target_path.exists():
            candidate = run_dir / f"{path.stem} (flat-root){path.suffix}"
            counter = 2
            while candidate.exists():
                candidate = run_dir / f"{path.stem} (flat-root {counter}){path.suffix}"
                counter += 1
            target_path = candidate

        path.rename(target_path)
        moved_messages.append(f"moved legacy file: {path.name} -> {target_path.name}")

    return moved_messages


def extract_article_from_base(description: object, title: object) -> str:
    candidates = [description, title]
    for value in candidates:
        if value is None or (isinstance(value, float) and pd.isna(value)):
            continue
        first_line = str(value).splitlines()[0].strip()
        if not first_line:
            continue
        if "," in first_line:
            candidate = first_line.rsplit(",", 1)[1].strip()
            candidate = candidate.strip("«»\"'“”")
            if candidate:
                return normalize_article(candidate)
        match = re.search(r"([A-Za-z0-9]+(?:[- ][A-Za-z0-9]+)*)$", first_line)
        if match:
            return normalize_article(match.group(1))
    return ""


def numeric_sort_key(path: Path) -> Tuple[int, object]:
    stem = path.stem.strip()
    if re.fullmatch(r"\d+", stem):
        return (0, int(stem))
    return (1, stem.lower())


def find_headers(ws) -> Dict[str, int]:
    headers: Dict[str, int] = {}
    for col_idx in range(1, ws.max_column + 1):
        value = ws.cell(2, col_idx).value
        if value is None:
            continue
        text = str(value).strip()
        if text:
            headers[text] = col_idx
    return headers


def ensure_headers(ws, header_map: Dict[str, int], headers: Iterable[str]) -> None:
    for header in headers:
        if header in header_map:
            continue
        target_col: Optional[int] = None
        for col_idx in range(1, ws.max_column + 1):
            value = ws.cell(2, col_idx).value
            if value is None or not str(value).strip():
                target_col = col_idx
                break
        if target_col is None:
            target_col = ws.max_column + 1
        ws.cell(2, target_col, header)
        header_map[header] = target_col


def iter_data_rows(ws, id_col: int) -> Iterable[Tuple[int, str]]:
    row = 5
    empty_streak = 0
    max_scan_rows = max(ws.max_row, 5) + 200

    while row <= max_scan_rows:
        raw_id = ws.cell(row, id_col).value
        normalized = normalize_id(raw_id)
        if normalized:
            empty_streak = 0
            yield row, normalized
        else:
            empty_streak += 1
            if empty_streak >= 200:
                break
        row += 1


def build_description(title: str, article: str, size_lines: List[str]) -> str:
    first_line = f"{title}, {article}".strip(", ")
    lines = [
        first_line if first_line else "!НЕТ!",
        "",
        "• Оригинал, приветствуются любые проверки",
        "• Новые, полный комплект",
        "• Все размеры в наличии в Москве",
        "",
        "Размеры :",
    ]

    if size_lines:
        lines.extend(size_lines)
    else:
        lines.append("• !НЕТ!")

    lines.extend(
        [
            "",
            "📍 Самовывоз с примеркой: Москва, м. Тульская – Даниловская мануфактура",
            "🚚 Доставка: Авито Доставка по всей России",
            "⚡️ Отправка в течение рабочего дня после оплаты",
            "___",
            "",
            "💬 Почему стоит купить у нас:",
            "• 100% оригинал",
            "• Лучшая цена на кроссовки в наличии",
            "• Быстро отвечаем и помогаем правильно подобрать размер с учетом особенностей модели",
            "___",
            "",
            "📩 Напишите в личные сообщения и мы с радостью ответим на любые ваши вопросы!",
        ]
    )
    return "\n".join(lines)


def is_kids_listing(goods_type: str, category: str) -> bool:
    goods = fold_text(clean_text(goods_type))
    cat = fold_text(clean_text(category))
    if "детск" in goods or "детск" in cat:
        return True
    return goods in {"для мальчиков", "для девочек"}


def build_size_field(size_ru: str, size_cm: str, kids: bool) -> str:
    if not size_ru:
        return "!НЕТ!"
    if kids and size_cm:
        return f"{size_ru} ({size_cm} см)"
    return size_ru


@dataclass
class AvitoRow:
    ad_id: str
    article: str
    title: str
    title_with_original: str
    brand: str
    price: str
    size_ru: str
    size_cm: str


class ImgbbUploader:
    def __init__(
        self,
        api_key: str,
        photos_root: Path,
        skip_upload: bool = False,
        max_consecutive_errors: int = 10,
        convert_heic_to_jpeg: bool = False,
        force_new_links: bool = False,
        newlink_tag: str = "",
    ):
        self.api_key = api_key
        self.photos_root = photos_root
        self.skip_upload = skip_upload
        self.max_consecutive_errors = max_consecutive_errors
        self.convert_heic_to_jpeg = convert_heic_to_jpeg
        self.force_new_links = force_new_links
        self.newlink_tag = clean_text(newlink_tag) or datetime.now().strftime("%Y%m%d%H%M%S")
        self._folder_map = self._build_folder_map()
        self._cache: Dict[str, Optional[str]] = {}
        self.logs: List[str] = []
        self._total_photos_to_process = 0
        self._processed_photo_files = 0
        self._last_progress_milestone = 0
        self._progress_step = 100
        self._consecutive_photo_errors = 0

    def _build_folder_map(self) -> Dict[str, Path]:
        mapping: Dict[str, Path] = {}
        if not self.photos_root.exists():
            return mapping

        for folder in sorted([p for p in self.photos_root.iterdir() if p.is_dir()], key=lambda x: x.name.lower()):
            names = {
                folder.name,
                folder.name.split("(", 1)[0].strip(),
            }
            for candidate in names:
                key = article_key(candidate)
                if key and key not in mapping:
                    mapping[key] = folder
        return mapping

    def _find_folder(self, article: str) -> Optional[Path]:
        return self._folder_map.get(article_key(article))

    def _prepare_image(self, file_path: Path, tmp_dir: Path) -> Path:
        if self.skip_upload:
            return file_path

        if not self.convert_heic_to_jpeg:
            return file_path

        if file_path.suffix.lower() not in HEIC_EXTENSIONS:
            return file_path

        return self._convert_heic_to_jpeg(file_path, tmp_dir)

    def _convert_heic_to_jpeg(self, file_path: Path, tmp_dir: Path) -> Path:
        out_path = tmp_dir / f"{file_path.stem}_{abs(hash(str(file_path)))}.jpg"
        command = [
            "sips",
            "-s",
            "format",
            "jpeg",
            "-s",
            "formatOptions",
            "best",
            str(file_path),
            "--out",
            str(out_path),
        ]
        result = subprocess.run(command, capture_output=True, text=True)
        if result.returncode != 0:
            raise RuntimeError(
                f"HEIC conversion failed for {file_path.name}: {result.stderr.strip() or result.stdout.strip()}"
            )
        return out_path

    def _decorate_link_for_newlink_mode(self, link: str, article: str, index: int) -> str:
        if not self.force_new_links:
            return link
        marker_raw = f"{self.newlink_tag}-{article_key(article)}-{index}"
        marker = quote_plus(marker_raw)
        separator = "&" if "?" in link else "?"
        return f"{link}{separator}newlink={marker}"

    def _upload_image(self, file_path: Path) -> str:
        if self.skip_upload:
            return f"mock://{file_path.name}"

        with file_path.open("rb") as image_file:
            response = requests.post(
                "https://api.imgbb.com/1/upload",
                params={"key": self.api_key},
                files={"image": image_file},
                timeout=120,
            )

        response.raise_for_status()
        data = response.json()
        if data.get("success"):
            return data["data"]["url"]
        raise RuntimeError(f"IMGBB upload error: {data.get('error') or data}")

    def links_for_article(self, article: str) -> Optional[str]:
        key = article_key(article)
        if key in self._cache:
            return self._cache[key]

        folder = self._find_folder(article)
        if folder is None:
            message = f"[PHOTO] Folder not found for article '{article}'"
            self.logs.append(message)
            print(f"⚠️ {message}")
            self._cache[key] = None
            return None

        files = [p for p in folder.iterdir() if p.is_file() and p.suffix.lower() in PHOTO_EXTENSIONS]
        files.sort(key=numeric_sort_key)
        if not files:
            message = f"[PHOTO] No photos in folder '{folder.name}'"
            self.logs.append(message)
            print(f"⚠️ {message}")
            self._cache[key] = None
            return None

        links: List[str] = []
        with tempfile.TemporaryDirectory() as tmp:
            tmp_dir = Path(tmp)
            for file_path in files:
                try:
                    prepared = self._prepare_image(file_path, tmp_dir)
                    try:
                        link = self._upload_image(prepared)
                    except requests.HTTPError as http_error:
                        # IMGBB may reject raw HEIC via API with 400. Fallback to JPEG conversion per-file.
                        needs_fallback = (
                            file_path.suffix.lower() in HEIC_EXTENSIONS
                            and not self.convert_heic_to_jpeg
                            and http_error.response is not None
                            and http_error.response.status_code == 400
                        )
                        if not needs_fallback:
                            raise
                        converted = self._convert_heic_to_jpeg(file_path, tmp_dir)
                        link = self._upload_image(converted)
                    if link:
                        link = self._decorate_link_for_newlink_mode(
                            link=link,
                            article=article,
                            index=len(links) + 1,
                        )
                        links.append(link)
                        self._consecutive_photo_errors = 0
                except Exception as error:
                    safe_error = mask_sensitive_text(error)
                    self.logs.append(f"[PHOTO] {article} | {folder.name} | {file_path.name} | {safe_error}")
                    self._consecutive_photo_errors += 1
                    print(
                        f"❌ Upload error {self._consecutive_photo_errors}"
                        + (
                            f"/{self.max_consecutive_errors}"
                            if self.max_consecutive_errors > 0
                            else ""
                        )
                        + f": folder='{folder.name}' file='{file_path.name}' | {safe_error}"
                    )
                    if self.max_consecutive_errors > 0 and self._consecutive_photo_errors >= self.max_consecutive_errors:
                        raise RuntimeError(
                            f"Reached {self._consecutive_photo_errors} consecutive photo upload errors. "
                            f"Stopping fail-fast on '{article}/{file_path.name}'."
                        )
                finally:
                    self._processed_photo_files += 1
                    if self._processed_photo_files // self._progress_step > self._last_progress_milestone:
                        self._last_progress_milestone = self._processed_photo_files // self._progress_step
                        if self._total_photos_to_process > 0:
                            print(
                                f"📊 Progress: {self._processed_photo_files}/{self._total_photos_to_process} photos processed"
                            )
                        else:
                            print(f"📊 Progress: {self._processed_photo_files} photos processed")

        if not links:
            self._cache[key] = None
            return None

        joined = "|".join(links)
        self._cache[key] = joined
        return joined

    def estimate_upload_scope(self, articles: Iterable[str]) -> Tuple[int, int]:
        folders_count = 0
        total_photos = 0

        for article in articles:
            folder = self._find_folder(article)
            if folder is None:
                continue

            files = [p for p in folder.iterdir() if p.is_file() and p.suffix.lower() in PHOTO_EXTENSIONS]
            if not files:
                continue

            folders_count += 1
            total_photos += len(files)

        return folders_count, total_photos

    def start_progress(self, total_photos: int, step: int = 100) -> None:
        self._total_photos_to_process = total_photos
        self._processed_photo_files = 0
        self._last_progress_milestone = 0
        self._progress_step = max(1, step)


def main() -> None:
    started_at = time.monotonic()
    parser = argparse.ArgumentParser(
        description="Builds new ready Avito file for ST_v2 with Base/AvitoNow mapping and IMGBB photo links."
    )
    parser.add_argument(
        "--skip-upload",
        action="store_true",
        help="Do not call IMGBB API. Writes mock links for quick local validation.",
    )
    parser.add_argument(
        "--max-consecutive-photo-errors",
        type=int,
        default=10,
        help="Fail-fast threshold for consecutive photo upload errors. Use 0 to disable fail-fast.",
    )
    parser.add_argument(
        "--newlink",
        nargs="?",
        const="next",
        metavar="vN",
        help=(
            "Force new link markers. "
            "Use '--newlink' for next version, or '--newlink v1' to write exactly into version v1."
        ),
    )
    args = parser.parse_args()
    if not args.skip_upload and not IMGBB_API_KEY:
        raise EnvironmentError("IMGBB_API_KEY is not set. Export it before running without --skip-upload.")

    migrated = migrate_legacy_output_files()
    for message in migrated:
        print(f"📦 {message}")

    avitonow_path = find_input_file([r"^Avitonow.*\.xlsx$"])
    base_path = find_input_file([r"^Base.*\.xlsx$"])
    ready_path = find_input_file([r"Готов.*файл.*\.xlsx$"])
    try:
        missing_base_path: Optional[Path] = find_input_file([r"Отсутств.*base.*\.xlsx$", r"missing.*base.*\.xlsx$"])
    except FileNotFoundError:
        missing_base_path = None

    date_str = datetime.now().strftime("%d.%m.%Y")
    newlink_requested = args.newlink is not None
    if newlink_requested and args.newlink not in (None, "next"):
        version = parse_version_arg(args.newlink)
    else:
        version = choose_next_version(date_str)
    run_output_dir = OUTPUT_ROOT_DIR / f"Готовые файлы {date_str} v{version}"
    run_output_dir.mkdir(parents=True, exist_ok=True)

    output_ready = run_output_dir / f"Готовый файл {date_str} v{version}.xlsx"
    output_photo_links = run_output_dir / f"IMGBB ссылки {date_str} v{version}.xlsx"
    output_log = run_output_dir / f"Лог {date_str} v{version}.txt"

    av_df = pd.read_excel(avitonow_path, sheet_name="Лист2")
    if ID_HEADER not in av_df.columns:
        raise KeyError(f"Column '{ID_HEADER}' not found in {avitonow_path.name}")

    av_by_id: Dict[str, AvitoRow] = {}
    article_size_rows: Dict[str, List[Tuple[str, str, str, str]]] = defaultdict(list)
    article_ru_sizes: Dict[str, set] = defaultdict(set)

    for _, row in av_df.iterrows():
        ad_id = normalize_id(row.get(ID_HEADER))
        if not ad_id:
            continue

        article = normalize_article(clean_text(row.get("Артикул")))
        title = normalize_human_case(row.get("Название"))
        title_with_original = add_original_suffix(title)
        brand = normalize_brand(row.get("Бренд"))
        price = get_discount_price_from_avitonow(row)
        size_ru = format_size_value(row.get("Размер: RU"))
        size_eu = format_size_value(row.get("Размер: EU"))
        size_us = format_size_value(row.get("Размер: US"))
        if not size_us:
            size_us = format_size_value(row.get("Размер US W"))
        size_cm = format_size_value(row.get("СМ (Длина стопы)"))

        av_by_id[ad_id] = AvitoRow(
            ad_id=ad_id,
            article=article,
            title=title,
            title_with_original=title_with_original,
            brand=brand,
            price=price,
            size_ru=size_ru,
            size_cm=size_cm,
        )

        article_size_rows[article_key(article)].append((size_us, size_eu, size_ru, size_cm))
        if size_ru:
            article_ru_sizes[article_key(article)].add(size_ru)

    article_has_multi_sizes = {key: len(values) > 1 for key, values in article_ru_sizes.items()}

    article_size_lines: Dict[str, List[str]] = {}
    for key, sizes in article_size_rows.items():
        unique_sizes = []
        seen = set()
        for size in sizes:
            if size in seen:
                continue
            seen.add(size)
            unique_sizes.append(size)

        def size_sort_key(item: Tuple[str, str, str, str]) -> Tuple[float, float, float]:
            def to_float(text: str) -> float:
                if not text:
                    return 9999.0
                try:
                    return float(text.replace(",", ".").split("-", 1)[0])
                except Exception:
                    return 9999.0

            return (to_float(item[1]), to_float(item[2]), to_float(item[3]))

        unique_sizes.sort(key=size_sort_key)
        lines = []
        for us, eu, ru, cm in unique_sizes:
            lines.append(f"• US {us or '!НЕТ!'} / EU {eu or '!НЕТ!'} / RU {ru or '!НЕТ!'} / {cm or '!НЕТ!'} см")
        article_size_lines[key] = lines

    base_df = pd.read_excel(base_path, sheet_name="Лист2", header=1)
    base_df = base_df[pd.to_numeric(base_df["Id"], errors="coerce").notna()].copy()

    base_grouped: Dict[str, Dict[str, List[str]]] = defaultdict(lambda: defaultdict(list))
    for _, row in base_df.iterrows():
        article = extract_article_from_base(row.get("Description"), row.get("Title"))
        key = article_key(article)
        if not key:
            continue
        for field in BASE_FIELDS:
            value = clean_text(row.get(field))
            if value:
                base_grouped[key][field].append(value)

    base_by_article: Dict[str, Dict[str, str]] = {}
    base_warnings: List[str] = []
    for key, values in base_grouped.items():
        resolved: Dict[str, str] = {}
        for field in BASE_FIELDS:
            variants = values.get(field, [])
            if not variants:
                resolved[field] = ""
                continue
            counter = Counter(variants)
            selected, _ = counter.most_common(1)[0]
            resolved[field] = selected
            if len(counter) > 1:
                base_warnings.append(
                    f"[BASE] {key}: field '{field}' has multiple values {dict(counter)}; selected '{selected}'"
                )
        base_by_article[key] = resolved

    fallback_by_article: Dict[str, Dict[str, str]] = {}
    if missing_base_path is not None:
        missing_df = pd.read_excel(missing_base_path)
        for _, row in missing_df.iterrows():
            article = normalize_article(row.get("Артикул"))
            key = article_key(article)
            if not key:
                continue
            fallback_by_article[key] = {
                "GoodsType": clean_text(row.get("Вид одежды")),
                "Color": clean_text(row.get("Цвет")),
                "ApparelType": clean_text(row.get("Вид одежды, обуви, аксессуаров")),
                "Category": clean_text(row.get("Категория")),
            }

    wb = load_workbook(ready_path)
    if TARGET_SHEET not in wb.sheetnames:
        raise KeyError(f"Sheet '{TARGET_SHEET}' not found in {ready_path.name}")
    ws = wb[TARGET_SHEET]

    header_map = find_headers(ws)
    required_headers = [
        ID_HEADER,
        NAME_HEADER,
        DESCRIPTION_HEADER,
        SIZE_HEADER,
        PHOTOS_HEADER,
        PRICE_HEADER,
        CATEGORY_HEADER,
        GOODS_TYPE_HEADER,
        CONDITION_HEADER,
        BRAND_HEADER,
        COLOR_HEADER,
        COLOR_MANUFACTURER_HEADER,
        APPAREL_TYPE_HEADER,
        MULTI_TITLE_HEADER,
    ]
    for header in required_headers:
        if header not in header_map:
            raise KeyError(f"Header '{header}' not found in row 2 of sheet '{TARGET_SHEET}'")
    ensure_headers(ws, header_map, [TYPE_GOODS_HEADER, TYPE_SHOES_HEADER])

    id_col = header_map[ID_HEADER]
    uploader = ImgbbUploader(
        api_key=IMGBB_API_KEY,
        photos_root=PHOTOS_ROOT,
        skip_upload=args.skip_upload,
        max_consecutive_errors=args.max_consecutive_photo_errors,
        convert_heic_to_jpeg=True,
        force_new_links=newlink_requested,
        newlink_tag=f"{date_str}-v{version}",
    )

    logs: List[str] = []
    photo_export_rows: List[Dict[str, str]] = []
    ready_rows = list(iter_data_rows(ws, id_col))
    processed_count = len(ready_rows)
    run_articles = []
    for _, ad_id in ready_rows:
        av_row = av_by_id.get(ad_id)
        if av_row is not None and av_row.article:
            run_articles.append(av_row.article)

    unique_run_articles = sorted(set(run_articles), key=lambda x: x.lower())
    folders_count, total_photos = uploader.estimate_upload_scope(unique_run_articles)
    uploader.start_progress(total_photos=total_photos, step=100)
    print(f"🚀 Start: folders with photos = {folders_count}, photos to process = {total_photos} 📸")

    for row_idx, ad_id in ready_rows:
        av_row = av_by_id.get(ad_id)

        if av_row is None:
            ws.cell(row_idx, header_map[NAME_HEADER], "!НЕТ!")
            ws.cell(row_idx, header_map[DESCRIPTION_HEADER], "!НЕТ!")
            ws.cell(row_idx, header_map[SIZE_HEADER], "!НЕТ!")
            ws.cell(row_idx, header_map[PHOTOS_HEADER], "!НЕТ!")
            ws.cell(row_idx, header_map[PRICE_HEADER], "!НЕТ!")
            ws.cell(row_idx, header_map[MULTI_TITLE_HEADER], "!НЕТ!")
            ws.cell(row_idx, header_map[GOODS_TYPE_HEADER], "!НЕТ!")
            ws.cell(row_idx, header_map[BRAND_HEADER], "!НЕТ!")
            ws.cell(row_idx, header_map[COLOR_HEADER], "!НЕТ!")
            ws.cell(row_idx, header_map[COLOR_MANUFACTURER_HEADER], "")
            ws.cell(row_idx, header_map[APPAREL_TYPE_HEADER], "!НЕТ!")
            ws.cell(row_idx, header_map[TYPE_GOODS_HEADER], "")
            ws.cell(row_idx, header_map[TYPE_SHOES_HEADER], "")
            photo_export_rows.append({ID_HEADER: ad_id, PHOTOS_HEADER: "!НЕТ!"})
            logs.append(f"[MATCH] ID '{ad_id}' missing in AvitoNow")
            continue

        article = av_row.article
        akey = article_key(article)

        ws.cell(row_idx, header_map[NAME_HEADER], av_row.title_with_original or "!НЕТ!")
        if article_has_multi_sizes.get(akey, False):
            multi_title = (
                sanitize_multi_title(av_row.title_with_original)
                or sanitize_multi_title(add_original_suffix(av_row.title))
                or sanitize_multi_title(article)
            )
            ws.cell(row_idx, header_map[MULTI_TITLE_HEADER], multi_title or "!НЕТ!")
            if not multi_title:
                logs.append(f"[MULTI] ID '{ad_id}' article '{article}' has empty sanitized multi title")
        else:
            ws.cell(row_idx, header_map[MULTI_TITLE_HEADER], "")
        ws.cell(row_idx, header_map[BRAND_HEADER], av_row.brand or "!НЕТ!")
        ws.cell(row_idx, header_map[COLOR_MANUFACTURER_HEADER], "")
        ws.cell(row_idx, header_map[PRICE_HEADER], av_row.price or "!НЕТ!")

        description = build_description(
            title=av_row.title,
            article=article,
            size_lines=article_size_lines.get(akey, []),
        )
        ws.cell(row_idx, header_map[DESCRIPTION_HEADER], description)

        fallback_values = fallback_by_article.get(akey)
        base_values = base_by_article.get(akey)
        goods_type = "!НЕТ!"
        color = "!НЕТ!"
        apparel_type = "!НЕТ!"
        if base_values is None:
            if fallback_values:
                goods_type = clean_text(fallback_values.get("GoodsType")) or "!НЕТ!"
                color = clean_text(fallback_values.get("Color")) or "!НЕТ!"
                apparel_type = clean_text(fallback_values.get("ApparelType")) or "!НЕТ!"
            else:
                logs.append(f"[BASE] Article '{article}' not found in Base and in fallback file")
        else:
            goods_type = clean_text(base_values.get("GoodsType")) or "!НЕТ!"
            color = clean_text(base_values.get("Color")) or "!НЕТ!"
            apparel_type = clean_text(base_values.get("ApparelType")) or "!НЕТ!"
            if fallback_values:
                goods_type = goods_type if goods_type != "!НЕТ!" else (clean_text(fallback_values.get("GoodsType")) or "!НЕТ!")
                color = color if color != "!НЕТ!" else (clean_text(fallback_values.get("Color")) or "!НЕТ!")
                apparel_type = apparel_type if apparel_type != "!НЕТ!" else (clean_text(fallback_values.get("ApparelType")) or "!НЕТ!")

            if goods_type == "!НЕТ!" or color == "!НЕТ!" or apparel_type == "!НЕТ!":
                logs.append(f"[BASE] Article '{article}' has incomplete Base values: {base_values}")

        ws.cell(row_idx, header_map[GOODS_TYPE_HEADER], goods_type)
        ws.cell(row_idx, header_map[COLOR_HEADER], color)
        ws.cell(row_idx, header_map[APPAREL_TYPE_HEADER], apparel_type)

        current_category = clean_text(ws.cell(row_idx, header_map[CATEGORY_HEADER]).value)
        category_value = current_category or DEFAULT_CATEGORY_VALUE
        if fallback_values:
            category_override = clean_text(fallback_values.get("Category"))
            if category_override:
                category_value = category_override
        ws.cell(row_idx, header_map[CATEGORY_HEADER], category_value)

        kids_listing = is_kids_listing(goods_type, category_value)
        ws.cell(row_idx, header_map[CONDITION_HEADER], "Новое" if kids_listing else "Новое с биркой")
        ws.cell(
            row_idx,
            header_map[SIZE_HEADER],
            build_size_field(av_row.size_ru, av_row.size_cm, kids_listing),
        )
        if not av_row.size_ru:
            logs.append(f"[SIZE] ID '{ad_id}' article '{article}' has empty RU size in AvitoNow")

        ws.cell(row_idx, header_map[TYPE_GOODS_HEADER], KIDS_TYPE_GOODS_VALUE if kids_listing else "")
        ws.cell(row_idx, header_map[TYPE_SHOES_HEADER], KIDS_TYPE_SHOES_VALUE if kids_listing else "")

        try:
            photo_links = uploader.links_for_article(article)
        except RuntimeError as fatal_photo_error:
            fatal_message = f"[FATAL] {fatal_photo_error}"
            logs.append(fatal_message)
            print(f"🛑 {fatal_message}")
            wb.save(output_ready)
            photo_df = pd.DataFrame(photo_export_rows, columns=[ID_HEADER, PHOTOS_HEADER])
            photo_df.to_excel(output_photo_links, index=False)

            elapsed_seconds = time.monotonic() - started_at
            full_logs = [
                f"Run time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                f"Mode: {'SKIP_UPLOAD' if args.skip_upload else 'FULL'}",
                f"Input AvitoNow: {avitonow_path.name}",
                f"Input Base: {base_path.name}",
                f"Input Ready: {ready_path.name}",
                f"Output Folder: {run_output_dir}",
                f"Output Ready: {output_ready}",
                f"Output Photo Links: {output_photo_links}",
                f"Processed rows before abort: {processed_count}",
                f"Abort reason: {fatal_photo_error}",
                "",
                "Runtime logs:",
            ]
            full_logs.extend(f"- {line}" for line in logs)
            if uploader.logs:
                full_logs.append("")
                full_logs.append("Uploader logs:")
                full_logs.extend(f"- {line}" for line in uploader.logs)
            full_logs.append("")
            full_logs.append(f"Elapsed seconds: {elapsed_seconds:.2f}")
            output_log.write_text("\n".join(full_logs), encoding="utf-8")

            raise SystemExit(2)

        if photo_links:
            ws.cell(row_idx, header_map[PHOTOS_HEADER], photo_links)
            photo_export_rows.append({ID_HEADER: ad_id, PHOTOS_HEADER: photo_links})
        else:
            ws.cell(row_idx, header_map[PHOTOS_HEADER], "!НЕТ!")
            photo_export_rows.append({ID_HEADER: ad_id, PHOTOS_HEADER: "!НЕТ!"})
            logs.append(f"[PHOTO] Article '{article}' has no uploaded links")

    wb.save(output_ready)

    photo_df = pd.DataFrame(photo_export_rows, columns=[ID_HEADER, PHOTOS_HEADER])
    photo_df.to_excel(output_photo_links, index=False)

    full_logs = []
    full_logs.append(f"Run time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    full_logs.append(f"Mode: {'SKIP_UPLOAD' if args.skip_upload else 'FULL'}")
    full_logs.append(f"Input AvitoNow: {avitonow_path.name}")
    full_logs.append(f"Input Base: {base_path.name}")
    full_logs.append(f"Input Ready: {ready_path.name}")
    full_logs.append(f"Output Folder: {run_output_dir}")
    full_logs.append(f"Output Ready: {output_ready}")
    full_logs.append(f"Output Photo Links: {output_photo_links}")
    full_logs.append(f"Processed rows: {processed_count}")
    full_logs.append(f"Base warnings: {len(base_warnings)}")
    full_logs.append(f"Runtime logs: {len(logs)}")
    full_logs.append(f"Uploader logs: {len(uploader.logs)}")
    full_logs.append("")

    if base_warnings:
        full_logs.append("Base warnings:")
        full_logs.extend(f"- {line}" for line in base_warnings)
        full_logs.append("")

    if logs:
        full_logs.append("Runtime logs:")
        full_logs.extend(f"- {line}" for line in logs)
        full_logs.append("")

    if uploader.logs:
        full_logs.append("Uploader logs:")
        full_logs.extend(f"- {line}" for line in uploader.logs)
        full_logs.append("")

    if not base_warnings and not logs and not uploader.logs:
        full_logs.append("All checks passed without warnings or errors.")

    elapsed_seconds = time.monotonic() - started_at
    full_logs.append("")
    full_logs.append(f"Elapsed seconds: {elapsed_seconds:.2f}")

    output_log.write_text("\n".join(full_logs), encoding="utf-8")

    print(f"✅ Done. Output file: {output_ready}")
    print(f"🖼️ Photo links file: {output_photo_links}")
    print(f"📝 Log file: {output_log}")
    print(f"⏱️ Elapsed: {elapsed_seconds:.2f} sec")


if __name__ == "__main__":
    main()
