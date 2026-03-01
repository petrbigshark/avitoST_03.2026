import importlib.util
import io
import re
import unittest
from contextlib import redirect_stdout
from pathlib import Path
from unittest.mock import patch

import pandas as pd
from openpyxl import Workbook


SCRIPT_PATH = Path(__file__).resolve().parents[1] / "build_ready_file.py"


def load_module():
    spec = importlib.util.spec_from_file_location("build_ready_file_e2e", SCRIPT_PATH)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module


def build_avitonow(path: Path) -> None:
    rows = [
        {
            "OSKELLY ID": 1,
            "Ссылка на OSKELLY": "oskelly.ru/products/1",
            "Фото": "",
            "Артикул": "ART-1",
            "Название": "adidas Y-3 GAZELLE white",
            "Категория": "Кроссовки и кеды",
            "Бренд": "aDIDAS",
            "Модель": "M1",
            "Размер: INT": "",
            "Размер: EU": 40,
            "Размер: RU": 39,
            "Размер: US": 7,
            "Размер US W": "",
            "СМ (Длина стопы)": 25,
            "Кол-во": 1,
            "Цена со скидкой": "10 000 ₽",
            "Цена до скидки": "15000",
            "Дефект": "Без дефекта",
            "Статус": "В продаже",
            "Уникальный идентификатор объявления": 1001,
        },
        {
            "OSKELLY ID": 2,
            "Ссылка на OSKELLY": "oskelly.ru/products/2",
            "Фото": "",
            "Артикул": "ART-1",
            "Название": "adidas Y-3 GAZELLE white",
            "Категория": "Кроссовки и кеды",
            "Бренд": "aDIDAS",
            "Модель": "M1",
            "Размер: INT": "",
            "Размер: EU": 41,
            "Размер: RU": 40,
            "Размер: US": 8,
            "Размер US W": "",
            "СМ (Длина стопы)": 26,
            "Кол-во": 1,
            "Цена со скидкой": "11 000 ₽",
            "Цена до скидки": "16000",
            "Дефект": "Без дефекта",
            "Статус": "В продаже",
            "Уникальный идентификатор объявления": 1002,
        },
        {
            "OSKELLY ID": 3,
            "Ссылка на OSKELLY": "oskelly.ru/products/3",
            "Фото": "",
            "Артикул": "ART-MISS",
            "Название": "nIKE sb dUNK low",
            "Категория": "Кроссовки и кеды",
            "Бренд": "nike x off-white",
            "Модель": "M2",
            "Размер: INT": "",
            "Размер: EU": 38,
            "Размер: RU": 37,
            "Размер: US": 6,
            "Размер US W": "",
            "СМ (Длина стопы)": 24.5,
            "Кол-во": 1,
            "Цена со скидкой": "12 000 ₽",
            "Цена до скидки": "17000",
            "Дефект": "Без дефекта",
            "Статус": "В продаже",
            "Уникальный идентификатор объявления": 1003,
        },
    ]
    pd.DataFrame(rows).to_excel(path, index=False, sheet_name="Лист2")


def build_base(path: Path) -> None:
    rows = [
        {
            "Id": 2001,
            "Title": "Model One",
            "Description": "Model One, ART-1\nline2",
            "GoodsType": "Женская обувь",
            "Color": "Синий",
            "ApparelType": "Кроссовки и кеды",
        }
    ]
    df = pd.DataFrame(rows)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Лист2", startrow=1, index=False)


def build_missing_base(path: Path) -> None:
    rows = [
        {
            "Артикул": "ART-MISS",
            "Вид одежды, обуви, аксессуаров": "Кроссовки и кеды",
            "Вид одежды": "Для мальчиков",
            "Цвет": "Красный",
            "Категория": "Детская одежда и обувь",
        }
    ]
    pd.DataFrame(rows).to_excel(path, index=False, sheet_name="Лист1")


def build_ready(path: Path, target_sheet: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = target_sheet
    ws.cell(1, 1, "Личные вещи - Одежда, обувь, аксессуары - Женская обувь - Кроссовки и кеды")

    headers = [
        "Уникальный идентификатор объявления",
        "Способ размещения",
        "Номер объявления на Авито",
        "Номер телефона",
        "Адрес",
        "Название объявления",
        "Описание объявления",
        "Ссылки на фото",
        "Способ связи",
        "Категория",
        "Цена",
        "Вид одежды",
        "Состояние",
        "Вид объявления",
        "Бренд одежды",
        "Цвет",
        "Цвет от производителя",
        "Материал основной части",
        "Соединять это объявление с другими объявлениями",
        "Название мультиобъявления",
        "Вид одежды, обуви, аксессуаров",
        "Размер",
        "Целевая аудитория",
        "AvitoDateEnd",
        "AvitoStatus",
        "Почта",
        "Название компании",
    ]
    for idx, header in enumerate(headers, start=1):
        ws.cell(2, idx, header)

    ws.cell(3, 1, "Обязательный")
    ws.cell(4, 1, "Подробнее о параметре")

    for row_idx, ad_id in enumerate([1001, 1002, 1003], start=5):
        ws.cell(row_idx, 1, ad_id)
        ws.cell(row_idx, 2, "Package")
        ws.cell(row_idx, 4, "79030042191")
        ws.cell(row_idx, 5, "Москва")
        ws.cell(row_idx, 9, "В сообщениях")
        ws.cell(row_idx, 10, "nan" if ad_id == 1002 else "Одежда, обувь, аксессуары")
        ws.cell(row_idx, 13, "Новое с биркой")
        ws.cell(row_idx, 14, "Товар приобретен на продажу")
        ws.cell(row_idx, 19, "Да")
        ws.cell(row_idx, 23, "Частные лица и бизнес")
        ws.cell(row_idx, 26, "mail@example.com")
        ws.cell(row_idx, 27, "sold twice")

    wb.save(path)


class BuildReadyE2ETests(unittest.TestCase):
    @staticmethod
    def _norm_numeric_text(value) -> str:
        text = str(value)
        if text.endswith(".0"):
            text = text[:-2]
        return text

    @staticmethod
    def _norm_price_text(value) -> str:
        text = str(value).strip()
        if re.fullmatch(r"\d+\.0+", text):
            return text.split(".", 1)[0]
        return "".join(ch for ch in text if ch.isdigit())

    def test_full_pipeline_skip_upload(self):
        mod = load_module()

        with self.subTest("prepare workspace"):
            tmp_root = Path(self._testMethodName)
            # Isolated temp dir inside test run directory
            test_dir = Path.cwd() / ".tmp_test_build_ready_e2e"
            if test_dir.exists():
                for item in sorted(test_dir.rglob("*"), reverse=True):
                    if item.is_file():
                        item.unlink()
                    else:
                        item.rmdir()
            test_dir.mkdir(parents=True, exist_ok=True)

            build_avitonow(test_dir / "Avitonow.xlsx")
            build_base(test_dir / "Base.xlsx")
            build_missing_base(test_dir / "Отсутствующие в base.xlsx")
            build_ready(test_dir / "Готовый файл.xlsx", mod.TARGET_SHEET)

            photos_root = test_dir / "Avito_ST_2"
            art1 = photos_root / "ART-1"
            art1.mkdir(parents=True, exist_ok=True)
            for i in range(1, 102):
                (art1 / f"{i}.jpg").write_bytes(b"jpg")

            art_missing = photos_root / "ART-MISS"
            art_missing.mkdir(parents=True, exist_ok=True)
            (art_missing / "1.jpg").write_bytes(b"jpg")

            mod.PROJECT_DIR = test_dir
            mod.PHOTOS_ROOT = photos_root
            mod.OUTPUT_ROOT_DIR = test_dir / "output Готовые файлы"
            mod.OUTPUT_ROOT_DIR.mkdir(parents=True, exist_ok=True)

        out = io.StringIO()
        with patch("sys.argv", ["build_ready_file.py", "--skip-upload"]), redirect_stdout(out):
            mod.main()
        stdout = out.getvalue()

        self.assertIn("🚀 Start:", stdout)
        self.assertIn("📊 Progress: 100/", stdout)
        self.assertIn("⏱️ Elapsed:", stdout)

        run_dirs = sorted(mod.OUTPUT_ROOT_DIR.glob("Готовые файлы * v*"))
        self.assertEqual(len(run_dirs), 1)
        run_dir = run_dirs[0]

        ready_files = list(run_dir.glob("Готовый файл *.xlsx"))
        links_files = list(run_dir.glob("IMGBB ссылки *.xlsx"))
        log_files = list(run_dir.glob("Лог *.txt"))
        self.assertEqual(len(ready_files), 1)
        self.assertEqual(len(links_files), 1)
        self.assertEqual(len(log_files), 1)

        ready_df = pd.read_excel(ready_files[0], sheet_name=mod.TARGET_SHEET, header=1)
        data = ready_df[ready_df[mod.ID_HEADER].astype(str).str.fullmatch(r"\d+")]
        by_id = {str(int(row[mod.ID_HEADER])): row for _, row in data.iterrows()}

        self.assertEqual(self._norm_numeric_text(by_id["1001"][mod.SIZE_HEADER]), "39")
        self.assertEqual(self._norm_numeric_text(by_id["1002"][mod.SIZE_HEADER]), "40")
        self.assertEqual(str(by_id["1003"][mod.SIZE_HEADER]), "37 (24,5 см)")
        self.assertEqual(self._norm_price_text(by_id["1001"][mod.PRICE_HEADER]), "10000")
        self.assertEqual(self._norm_price_text(by_id["1002"][mod.PRICE_HEADER]), "11000")
        self.assertEqual(self._norm_price_text(by_id["1003"][mod.PRICE_HEADER]), "12000")

        description = str(by_id["1001"][mod.DESCRIPTION_HEADER])
        self.assertIn("Adidas Y-3 Gazelle White, ART-1", description)
        self.assertNotIn("Adidas Y-3 Gazelle White Оригинал, ART-1", description)
        self.assertIn("RU 39", description)
        self.assertIn("RU 40", description)

        self.assertEqual(str(by_id["1001"][mod.NAME_HEADER]), "Adidas Y-3 Gazelle White Оригинал")
        self.assertEqual(str(by_id["1001"][mod.BRAND_HEADER]), "Adidas")
        self.assertEqual(str(by_id["1001"][mod.MULTI_TITLE_HEADER]), "Adidas Y-3 Gazelle White Оригинал")
        self.assertEqual(str(by_id["1002"][mod.MULTI_TITLE_HEADER]), "Adidas Y-3 Gazelle White Оригинал")
        self.assertTrue(pd.isna(by_id["1001"][mod.COLOR_MANUFACTURER_HEADER]) or str(by_id["1001"][mod.COLOR_MANUFACTURER_HEADER]).strip() == "")
        self.assertEqual(str(by_id["1002"][mod.CATEGORY_HEADER]), "Одежда, обувь, аксессуары")

        self.assertEqual(str(by_id["1003"][mod.CATEGORY_HEADER]), "Детская одежда и обувь")
        self.assertEqual(str(by_id["1003"][mod.GOODS_TYPE_HEADER]), "Для мальчиков")
        self.assertEqual(str(by_id["1003"][mod.COLOR_HEADER]), "Красный")
        self.assertEqual(str(by_id["1003"][mod.APPAREL_TYPE_HEADER]), "Кроссовки и кеды")
        self.assertEqual(str(by_id["1003"][mod.BRAND_HEADER]), "Nike")
        self.assertEqual(str(by_id["1003"][mod.CONDITION_HEADER]), "Новое")
        self.assertTrue(pd.isna(by_id["1003"][mod.MULTI_TITLE_HEADER]) or str(by_id["1003"][mod.MULTI_TITLE_HEADER]).strip() == "")
        self.assertIn(mod.TYPE_GOODS_HEADER, ready_df.columns)
        self.assertIn(mod.TYPE_SHOES_HEADER, ready_df.columns)
        self.assertEqual(str(by_id["1003"][mod.TYPE_GOODS_HEADER]), "Обувь")
        self.assertEqual(str(by_id["1003"][mod.TYPE_SHOES_HEADER]), "Кроссовки")

        links_df = pd.read_excel(links_files[0])
        links_map = {str(int(row[mod.ID_HEADER])): str(row[mod.PHOTOS_HEADER]) for _, row in links_df.iterrows()}
        self.assertNotEqual(links_map["1001"], "!НЕТ!")
        self.assertNotEqual(links_map["1003"], "!НЕТ!")
        self.assertGreaterEqual(links_map["1001"].count("|"), 100)
        self.assertIn("mock://", links_map["1001"])


if __name__ == "__main__":
    unittest.main()
